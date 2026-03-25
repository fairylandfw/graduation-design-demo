from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from app.models.models import Article
from analyzer.sentiment_analyzer import SentimentAnalyzer
from analyzer.topic_analyzer import TopicAnalyzer
from config.config import Config

class ReportGenerator:
    """报告生成器"""

    def __init__(self):
        self.sentiment_analyzer = SentimentAnalyzer(Config.DICT_DIR)
        self.topic_analyzer = TopicAnalyzer(Config.STOPWORDS_PATH)

    def generate_excel_report(self, start_date, end_date, output_path):
        """生成Excel报告"""
        # 查询数据
        articles = Article.query.filter(
            Article.publish_time >= start_date,
            Article.publish_time <= end_date
        ).all()

        # 创建工作簿
        wb = Workbook()

        # 概览页
        ws_overview = wb.active
        ws_overview.title = '概览'
        self._write_overview(ws_overview, articles, start_date, end_date)

        # 情感分析页
        ws_sentiment = wb.create_sheet('情感分析')
        self._write_sentiment_analysis(ws_sentiment, articles)

        # 热点话题页
        ws_topics = wb.create_sheet('热点话题')
        self._write_topics(ws_topics, articles)

        # 保存
        wb.save(output_path)
        return output_path

    def _write_overview(self, ws, articles, start_date, end_date):
        """写入概览信息"""
        ws['A1'] = '网络舆情分析报告'
        ws['A1'].font = Font(size=16, bold=True)

        ws['A3'] = '报告时间范围：'
        ws['B3'] = f'{start_date.strftime("%Y-%m-%d")} 至 {end_date.strftime("%Y-%m-%d")}'

        ws['A4'] = '数据总量：'
        ws['B4'] = len(articles)

        # 情感统计
        sentiment_count = {'positive': 0, 'neutral': 0, 'negative': 0}
        for article in articles:
            if article.sentiment:
                sentiment_count[article.sentiment] += 1

        ws['A6'] = '情感分布：'
        ws['A7'] = '正面：'
        ws['B7'] = sentiment_count['positive']
        ws['A8'] = '中性：'
        ws['B8'] = sentiment_count['neutral']
        ws['A9'] = '负面：'
        ws['B9'] = sentiment_count['negative']

    def _write_sentiment_analysis(self, ws, articles):
        """写入情感分析详情"""
        headers = ['标题', '来源', '发布时间', '情感', '情感得分']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)

        for row, article in enumerate(articles, 2):
            ws.cell(row, 1, article.title)
            ws.cell(row, 2, article.source)
            ws.cell(row, 3, article.publish_time.strftime('%Y-%m-%d %H:%M') if article.publish_time else '')
            ws.cell(row, 4, article.sentiment or '')
            ws.cell(row, 5, article.sentiment_score or '')

    def _write_topics(self, ws, articles):
        """写入热点话题"""
        texts = [a.content for a in articles if a.content]
        hotwords = self.topic_analyzer.extract_hotwords(texts, topK=30)

        headers = ['排名', '关键词', '出现次数']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)

        for row, (idx, hotword) in enumerate(zip(range(1, len(hotwords)+1), hotwords), 2):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, hotword['word'])
            ws.cell(row, 3, hotword['count'])
